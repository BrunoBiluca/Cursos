import openpyxl as xl
from openpyxl.chart import Reference, BarChart


wb = xl.load_workbook('transactions.xlsx')
sheet = wb["Sheet1"]

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    cell_corrected = sheet.cell(row, 4)
    cell_corrected.value = corrected_price

values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'E2')

wb.save('transactions.xlsx')
